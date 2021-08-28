#!/usr/bin/env python3
import smbclient
from datetime import datetime, date, time, timedelta, tzinfo
from openpyxl import load_workbook
import csv
from pytz import timezone
from itertools import zip_longest

import argparse

CHANNEL = "RedeMetropole"
ARCHIVE_AVAILABLE = 0
AVAILABLE = 1
LOGO_URL = "https://assets.iptv2022.com/uploads/channel/11208/Rede_Metropole__pt_.jpeg"

parser = argparse.ArgumentParser(description=f'Get program for {CHANNEL}')
parser.add_argument('--output', default="out.csv", help='output csv name')

sao_paulo = timezone('America/Sao_Paulo')

weekdays = {
    'QUINTA-FEIRA': 4,
    'DOMINGO': 7,
    'SÁBADO': 6,
    'QUARTA-FEIRA': 3,
    'SEGUNDA-FEIRA': 1,
    'SEXTA-FEIRA': 5,
    'TERÇA-FEIRA': 2
}

template = {
    'channel': CHANNEL,
    'logo': LOGO_URL,
    'archive_available': ARCHIVE_AVAILABLE,
    'available': AVAILABLE
}

def main():
    (year, week, _) = datetime.now().isocalendar()

    smbclient.ClientConfig(username='user10', password='ec52cd58')
    args = parser.parse_args()
    data = [list() for _ in range(7)]

    with smbclient.open_file('\\194.35.48.33\home\RedeMetropole\{}.xlsx'.format(year), mode="rb") as f:
        wb = load_workbook(f, read_only=True)
        sheet = wb.worksheets[0]
        i = 0
        for row in sheet.iter_rows(max_col=4, values_only=True):
            if not row[0] or row[0] == 'HORÁRIO INÍCIO':
                continue

            if isinstance(row[0], str) and row[0].strip() in weekdays:
                i = weekdays[row[0].strip()]
                print(i)
                continue

            if not isinstance(row[0], time):
                t = datetime.strptime(row[0].strip(), "%H:%M:%S").time()
            else:
                t = row[0]

            data[i-1].append({
                "start_at": t,
                "title": row[1],
                "description": row[3]
            })

    with open(args.output, 'w', newline='') as csvfile:
        fieldnames = ['start_at', 'end_at', 'channel', 'title', 'logo', 'description', "archive_available", "available"]
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        for day, data in enumerate(data, 1):
            d = date.fromisocalendar(year, week, day)
            for i, next_i in zip_longest(data, data[1:]):
                if next_i:
                    i["end_at"] = next_i["start_at"]
                else:
                    i["end_at"] = time(hour = 23, minute = 59)

                for k in "start_at", "end_at":
                    i[k] = sao_paulo.localize(datetime.combine(d, i[k])).isoformat()

                writer.writerow({**template, **i})


if __name__ == '__main__':
    main()
