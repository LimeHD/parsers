#!/usr/bin/env python3
import smbclient
from datetime import datetime, date, time, timedelta, tzinfo
from openpyxl import load_workbook
import csv
import locale
import re

import argparse

locale.setlocale(locale.LC_TIME, "es_ES.UTF-8")

CHANNEL = "AlientoVisiónHFN"
ARCHIVE_AVAILABLE = 0
AVAILABLE = 1
LOGO_URL = "https://assets.iptv2022.com/uploads/channel/10893/thumb_Aliento_Vision.png"

parser = argparse.ArgumentParser(description=f'Get program for {CHANNEL}')
parser.add_argument('--output', default="out.csv", help='output csv name')

weekday = {
    'DOMINGO': 7,
    'JUEVES': 4,
    'LUNES': 1,
    'MARTES': 2,
    'MIERCOLES': 3,
    'SABADO': 6,
    'VIERNES': 5
}

class EST(tzinfo):
    def utcoffset(self, dt):
        return timedelta(hours = -5)

    def tzname(self, dt):
        return "EST"

    def dst(self, dt):
        return timedelta(0)

def main():
    (year, week, _) = datetime.now().isocalendar()

    smbclient.ClientConfig(username='user1', password='5f217c95')
    args = parser.parse_args()

    with smbclient.open_file('\\194.35.48.33\home\AlientoVisiónHFN\{0}_{1}.xlsx'.format(year, week), mode="rb") as f:
        wb = load_workbook(f, read_only=True)
        with open(args.output, 'w', newline='') as csvfile:
            fieldnames = ['start_at', 'end_at', 'channel', 'title', 'logo', 'description', "archive_available", "available"]
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            for sheet in wb:
                for row in sheet.iter_rows(min_row=3, max_col=5, values_only=True):
                    d = date.fromisocalendar(year, week, weekday[row[0]])
                    if isinstance(row[1], str):
                        s = re.match('(\d{2}:\d{2}:\d{2} [AP]M) .*', row[1]).group(1)
                        t = datetime.strptime(s, "%I:%M:%S %p").time()
                    else:
                        t = row[1].time()

                    start_time = datetime.combine(d, t).replace(tzinfo=EST())
                    if isinstance(row[2], str):
                        diff = timedelta(seconds = int(re.match("(\d+)[mM]", row[2]).group(1))*60)
                    else:
                        diff = datetime.combine(datetime.min, row[2]) - datetime.min
                    end_time = start_time + diff

                    writer.writerow({
                        'start_at': start_time.isoformat(),
                        'end_at': end_time.isoformat(),
                        'channel': CHANNEL,
                        'title': row[3],
                        'logo': LOGO_URL,
                        'description': row[4],
                        'archive_available': ARCHIVE_AVAILABLE,
                        'available': AVAILABLE
                    })

if __name__ == '__main__':
    main()
