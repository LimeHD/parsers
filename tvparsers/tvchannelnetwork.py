#!/usr/bin/env python3
import smbclient
from datetime import datetime, date, time, timedelta, tzinfo
from openpyxl import load_workbook
import csv
from pytz import timezone
from itertools import zip_longest

import argparse

CHANNEL = "TVChannelNetwork"
ARCHIVE_AVAILABLE = 0
AVAILABLE = 1
LOGO_URL = "https://assets.iptv2022.com/uploads/channel/10921/thumb_photo_2021-05-28_17-29-40.jpg"

parser = argparse.ArgumentParser(description=f'Get program for {CHANNEL}')
parser.add_argument('--output', default="out.csv", help='output csv name')

sao_paulo = timezone('America/Sao_Paulo')

template = {
    'channel': CHANNEL,
    'logo': LOGO_URL,
    'description': "",
    'archive_available': ARCHIVE_AVAILABLE,
    'available': AVAILABLE
}

def main():
    (year, week, _) = datetime.now().isocalendar()

    smbclient.ClientConfig(username='user10', password='ec52cd58')
    args = parser.parse_args()

    with smbclient.open_file('\\194.35.48.33\home\TVChannelNetwork\{}.xlsx'.format(year), mode="rb") as f:
        wb = load_workbook(f, read_only=True)
        sheet = wb.worksheets[0]
        times = [row[0] for row in sheet.iter_rows(min_row = 2, min_col = 1, max_col = 1, values_only = True)]
        times.append(time(hour = 23, minute = 59))
        timeslots = list(zip(times, times[1:]))

        with open(args.output, 'w', newline='') as csvfile:
            fieldnames = ['start_at', 'end_at', 'channel', 'title', 'logo', 'description', "archive_available", "available"]
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            for i in range(2,9):
                d = date.fromisocalendar(year, week, i-1)
                for (t1, t2), row in zip(timeslots, sheet.iter_rows(min_row = 2, min_col = i, max_col = i, values_only = True)):
                    writer.writerow({**template, **{
                        "start_at": sao_paulo.localize(datetime.combine(d, t1)).isoformat(),
                        "end_at": sao_paulo.localize(datetime.combine(d, t2)).isoformat(),
                        "title": row[0]
                    }})

if __name__ == '__main__':
    main()
