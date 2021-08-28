#!/usr/bin/env python3
import smbclient
from datetime import datetime, date, time, timedelta, tzinfo
from openpyxl import load_workbook
import csv
from pytz import timezone
from itertools import zip_longest

import argparse

CHANNEL = "Peruchannel"
ARCHIVE_AVAILABLE = 0
AVAILABLE = 1
LOGO_URL = "https://assets.iptv2022.com/uploads/channel/11278/thumb_Peru%CC%81_Channel__es_.jpeg"

parser = argparse.ArgumentParser(description=f'Get program for {CHANNEL}')
parser.add_argument('--output', default="out.csv", help='output csv name')

lima = timezone('America/Lima')

template = {
    'channel': CHANNEL,
    'logo': LOGO_URL,
    'description': "",
    'archive_available': ARCHIVE_AVAILABLE,
    'available': AVAILABLE
}

def main():
    (year, week, _) = datetime.now().isocalendar()

    smbclient.ClientConfig(username='user10', password='ec52cd58')
    args = parser.parse_args()
    weekday = []
    saturday = []
    sunday = []

    with smbclient.open_file('\\194.35.48.33\home\Peruchannel\{}.xlsx'.format(year), mode="rb") as f:
        wb = load_workbook(f, read_only=True)
        for sheet in wb:
            l = weekday
            for row in sheet.iter_rows(max_col=3, values_only=True):
                if row[0] == "Sábados":
                    l = saturday
                if row[0] == "Domingos":
                    l = sunday
                l.append(row[1:])

    for i in weekday, saturday, sunday:
        i.append((time(hour = 23, minute = 59), ""))

    with open(args.output, 'w', newline='') as csvfile:
        fieldnames = ['start_at', 'end_at', 'channel', 'title', 'logo', 'description', "archive_available", "available"]
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        for weekdays, data in (range(1, 6), weekday), ([6], saturday), ([7], sunday):
            for weekday in weekdays:
                for i, next_i in zip(data, data[1:]):
                    d = date.fromisocalendar(year, week, weekday)

                    writer.writerow({**template, **{
                        "start_at": lima.localize(datetime.combine(d, i[0])).isoformat(),
                        "end_at": lima.localize(datetime.combine(d, next_i[0])).isoformat(),
                        "title": i[1]
                    }})


if __name__ == '__main__':
    main()
