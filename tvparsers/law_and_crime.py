#!/usr/bin/env python3
import smbclient
from datetime import datetime, date, timedelta, tzinfo
from openpyxl import load_workbook
import csv

import argparse

CHANNEL = "Law&Crime"
ARCHIVE_AVAILABLE = 0
AVAILABLE = 1
LOGO_URL = "https://assets.iptv2022.com/uploads/channel/10760/thumb_photo_2021-05-06_08.jpeg"

parser = argparse.ArgumentParser(description=f'Get program for {CHANNEL}')
parser.add_argument('--output', default="out.csv", help='output csv name')

weekday = {
    "Monday": 1,
    "Tuesday": 2,
    "Wednesday": 3,
    "Thursday": 4,
    "Friday": 5,
    "Saturday": 6,
    "Sunday": 7
}

class EST(tzinfo):
    def utcoffset(self, dt):
        return timedelta(hours = -5)

    def tzname(self, dt):
        return "EST"

    def dst(self, dt):
        return timedelta(0)

def main():
    (year, week, _) = datetime.now().isocalendar()

    smbclient.ClientConfig(username='user1', password='5f217c95')
    args = parser.parse_args()

    with smbclient.open_file('\\194.35.48.33\home\Law&Crime\{0}_{1}.xlsx'.format(year, week), mode="rb") as f:
        wb = load_workbook(f, read_only=True)
        with open(args.output, 'w', newline='') as csvfile:
            fieldnames = ['start_at', 'end_at', 'channel', 'title', 'logo', 'description', "archive_available", "available"]
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

            for sheet in wb:
                wd = weekday[sheet.title.strip()]
                d = date.fromisocalendar(year, week, wd)

                for row in sheet.iter_rows(min_row=2, min_col=3, max_col=8):
                    if row[0].data_type != 'd':
                        continue

                    writer.writerow({
                        'start_at': datetime.combine(d, row[0].value).replace(tzinfo=EST()).isoformat(),
                        'end_at': datetime.combine(d, row[1].value).replace(tzinfo=EST()).isoformat(),
                        'channel': CHANNEL,
                        'title': row[2].value,
                        'logo': LOGO_URL,
                        'description': row[5].value,
                        'archive_available': ARCHIVE_AVAILABLE,
                        'available': AVAILABLE
                    })

if __name__ == '__main__':
    main()
