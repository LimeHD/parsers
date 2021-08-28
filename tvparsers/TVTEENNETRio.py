#!/usr/bin/env python3
import smbclient
from datetime import datetime, date, time, timedelta, tzinfo
from openpyxl import load_workbook
import csv
from pytz import timezone
import re
from itertools import zip_longest

import argparse

CHANNEL = "TVTEENNETRio"
ARCHIVE_AVAILABLE = 0
AVAILABLE = 1
LOGO_URL = "https://assets.iptv2022.com/uploads/channel/11082/thumb_TVTEENNET-RIO_-_AGRUPADO-2021.png"

parser = argparse.ArgumentParser(description=f'Get program for {CHANNEL}')
parser.add_argument('--output', default="out.csv", help='output csv name')

sao_paulo = timezone('America/Sao_Paulo')
regexp = re.compile('(\d{1,2}:\d{2}) - (.*)')

template = {
    'channel': CHANNEL,
    'logo': LOGO_URL,
    'description': "",
    'archive_available': ARCHIVE_AVAILABLE,
    'available': AVAILABLE
}

def main():
    (year, week, _) = datetime.now().isocalendar()

    smbclient.ClientConfig(username='user10', password='ec52cd58')
    args = parser.parse_args()
    data = []

    with smbclient.open_file('\\194.35.48.33\home\TVTEENNETRio\{}.xlsx'.format(year), mode="rb") as f:
        wb = load_workbook(f, read_only=True)
        for sheet in wb:
            for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
                if not row[0]:
                    continue
                m = regexp.match(row[0])
                if not m:
                    continue

                data.append(m.groups())

    data.append(("23:59", "")) # stub for last program end

    with open(args.output, 'w', newline='') as csvfile:
        fieldnames = ['start_at', 'end_at', 'channel', 'title', 'logo', 'description', "archive_available", "available"]
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        for weekday in range(1, 8):
            for i, next_i in zip(data, data[1:]):
                d = date.fromisocalendar(year, week, weekday)

                writer.writerow({**template, **{
                    "start_at": sao_paulo.localize(datetime.combine(d, datetime.strptime(i[0], "%H:%M").time())).isoformat(),
                    "end_at": sao_paulo.localize(datetime.combine(d, datetime.strptime(next_i[0], "%H:%M").time())).isoformat(),
                    "title": i[1]
                }})


if __name__ == '__main__':
    main()
