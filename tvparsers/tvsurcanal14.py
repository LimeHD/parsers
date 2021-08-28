#!/usr/bin/env python3
import smbclient
from datetime import datetime, date, time, timedelta, tzinfo
from openpyxl import load_workbook
import csv
from pytz import timezone
from itertools import zip_longest

import argparse

CHANNEL = "TvSurCanal14"
ARCHIVE_AVAILABLE = 0
AVAILABLE = 1
LOGO_URL = "https://assets.iptv2022.com/uploads/channel/11544/thumb_TV_Sur_Canal_14__es_.jpeg"

parser = argparse.ArgumentParser(description=f'Get program for {CHANNEL}')
parser.add_argument('--output', default="out.csv", help='output csv name')

costa_rica = timezone('America/Costa_Rica')

def main():
    (year, week, _) = datetime.now().isocalendar()

    smbclient.ClientConfig(username='user1', password='5f217c95')
    args = parser.parse_args()
    rows = []

    with smbclient.open_file('\\194.35.48.33\home\TvSurCanal14\{0}_{1}.xlsx'.format(year, week), mode="rb") as f:
        wb = load_workbook(f, read_only=True)
        for sheet in wb:
            for row in sheet.iter_rows(min_row=3, max_col=3, values_only=True):
                d1 = date.fromisocalendar(year, week, 6)
                d2 = date.fromisocalendar(year, week, 7)
                for (d, i) in ((d1, 1), (d2, 2)):
                    start_time = costa_rica.localize(datetime.combine(d, row[0]))
                    rows.append({
                        'start_at': start_time,
                        'channel': CHANNEL,
                        'title': row[i],
                        'logo': LOGO_URL,
                        'description': "",
                        'archive_available': ARCHIVE_AVAILABLE,
                        'available': AVAILABLE
                    })

    rows.sort(key=lambda x:x["start_at"])

    with open(args.output, 'w', newline='') as csvfile:
        fieldnames = ['start_at', 'end_at', 'channel', 'title', 'logo', 'description', "archive_available", "available"]
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        for row, next_row in zip_longest(rows, rows[1:]):
            row["end_at"] = next_row["start_at"] if next_row else datetime.combine(row["start_at"].date(), time())
            row["start_at"] = row["start_at"].isoformat()
            row["end_at"] = row["end_at"].isoformat()
            writer.writerow(row)

if __name__ == '__main__':
    main()
